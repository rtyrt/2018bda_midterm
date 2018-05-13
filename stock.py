from openpyxl import load_workbook
import operator
import re
import string
import collections
import math
import pickle
from datetime import datetime
#————————————————————————FUNCTIONS——————————————————————————

#input:(name,table)=(stock name string, excel sheet)
#output:dictionary in the form of{date time:price}
#
def get_stock_data(name,table_2016,table_2017):
    stock={}
    for row in table_2016:
        if (row[0].value==name):
            change_date_format=row[1].value.strftime('%Y-%m-%d')
            stock[change_date_format]=[]
            stock[change_date_format].append(row[2].value)
    for row in table_2017:
        if (row[0].value==name):
            change_date_format=row[1].value.replace("/","-")
            stock[change_date_format]=[]
            stock[change_date_format].append(row[2].value)

    filename = "./stock_data_"+name.split(" ")[0]+".p"
    fread = open(filename, "wb")
    pickle.dump(stock, fread)
    fread.close()
    return stock

#input:(name)=dictionary in the form of{date time:price}
#output: a list of daily_return
#
def get_return(name):   
    price=list(name.values())
    daily_return=[]
    for i in range(len(price)):
        if i<(len(price)-1):
            r=(price[i][0]-price[i+1][0])/price[i+1][0]
            daily_return.append(r)

    return daily_return


#input: a single list of daily_return
#output: a list of 3 element[25%,50%,75%]百分位數
#
#def return_para(return_list):
 #   rr_list=np.array(return_list)
  #  median=np.percentile(rr_list, 50)
  #   third_quartile=np.percentile(rr_list,75)
  #  first_quartile=np.percentile(rr_list,25)
    
   # para=[first_quartile,median,third_quartile]
    #return para


#input:(return_list,name)=(a single list of daily_return,dictionary in the form of{date time:price})
#output: dictionary in the form of{datetime: [price, 1/0]}
#
def assign_return_value(return_list,name):
    for i in range(len(return_list)):
        if (return_list[i]>0.0054):
            return_list[i]=1
        else:
            return_list[i]=0
    
    for key,each_daily_return in zip(name,return_list):
        name[key].append(each_daily_return)
        
    return name

#input:(name,table)=(stock name string, excel sheet)
#output:dictionary in the form of{date time:price}
#
def load_origin_stock_data(name):
    stock={}
    filename = "./stock_data_"+name.split(" ")[0]+".p"
    fileObject = open(filename,'rb')
    stock = pickle.load(fileObject)
    return stock

#input:(name,table)=(stock name string, excel sheet)
#output:dictionary in the form of{date time:price}
#
def load_stock_data(name):
    stock={}
    filename = "./data_"+name+".p"
    fileObject = open(filename,'rb')
    stock = pickle.load(fileObject)
    return stock


#--------------------Main Code------------------------------------

# #load excel file 2016 stock 
# wb = load_workbook(filename=r'2016_stock_data.xlsx')
# ws = wb.worksheets[0]
# table_row=ws.rows

# #load excel file 2017 stock 
# wb_2017 = load_workbook(filename=r'2017_stock_data.xlsx')
# ws_2017 = wb_2017.worksheets[0]
# table_row_2017=ws_2017.rows


# #filter out targeted stock info
# asus=get_stock_data("2357 華碩",table_row,table_row_2017)
# asus=load_origin_stock_data("2357 華碩")

# acer=get_stock_data("2353 宏碁",table_row,table_row_2017)
# acer=load_origin_stock_data("2353 宏碁")

# #calculate return for each stock
# stock_asus_return=get_return(asus)
# stock_acer_return=get_return(acer)

# #assign value 1/0/-1 to daily return
# stock_asus=assign_return_value(stock_asus_return,asus)
# stock_acer=assign_return_value(stock_acer_return,acer)

# #save as pickle file
# filename = "./data_asus.p"
# fread = open(filename, "wb")
# pickle.dump(stock_asus, fread)
# fread.close()

# filename = "./data_acer.p"
# fread = open(filename, "wb")
# pickle.dump(stock_acer, fread)
# fread.close()
